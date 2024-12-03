from flask import Flask, render_template, request, send_file, jsonify
from dotenv import load_dotenv
from openai import OpenAI
import os
import openpyxl
from datetime import datetime
import json

# Charger les variables d'environnement
load_dotenv()

# Configuration
app = Flask(__name__)
client = OpenAI(
    api_key=os.getenv('OPENAI_API_KEY'),
    base_url="https://api.openai.com/v1"  # Assurez-vous d'utiliser la bonne URL de base
)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate_excel():
    try:
        prompt = request.json['prompt']
        
        # Appel à l'API OpenAI avec la nouvelle syntaxe
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Tu es un assistant spécialisé dans la création de données pour Excel. Réponds uniquement avec du JSON valide."},
                {"role": "user", "content": f"Crée des données Excel pour: {prompt}. Réponds avec un JSON contenant: 1) headers (les en-têtes des colonnes) 2) data (les données pour chaque ligne). Limite à 10 lignes maximum."}
            ]
        )

        # Extraire et parser la réponse
        content = response.choices[0].message.content
        excel_data = json.loads(content)

        # Créer le fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active

        # Ajouter les en-têtes
        for col, header in enumerate(excel_data['headers'], 1):
            ws.cell(row=1, column=col, value=header)

        # Ajouter les données
        for row, data_row in enumerate(excel_data['data'], 2):
            for col, value in enumerate(data_row, 1):
                ws.cell(row=row, column=col, value=value)

        # Sauvegarder le fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"generated_excel_{timestamp}.xlsx"
        filepath = os.path.join("static", "downloads", filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)

        return jsonify({
            "success": True,
            "filename": filename
        })

    except Exception as e:
        print(f"Error: {str(e)}")  # Ajout d'un log pour le débogage
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

@app.route('/download/<filename>')
def download_file(filename):
    return send_file(
        os.path.join("static", "downloads", filename),
        as_attachment=True
    )

if __name__ == '__main__':
    app.run(debug=True)
